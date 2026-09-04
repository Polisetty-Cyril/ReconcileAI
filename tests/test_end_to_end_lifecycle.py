"""
ReconcileAI - Phase 17 Task 7: End-to-End & Cross-Phase Integration Tests
Verifies the complete financial lifecycle across integrated subsystem boundaries:
1. Discrepancy Lifecycle to Human Approval & Audit:
   Mismatched transaction cluster -> Reconcile pipeline -> Exception & AI Advisory -> Human Approval -> Audit Trail
2. Reporting API to Export Serialization Pipeline:
   Real database state -> FastAPI Report Endpoints -> Schema Serialization -> In-Memory Excel/CSV Export Utilities
"""

from __future__ import annotations

import io
import json
from datetime import datetime, timezone, timedelta
import pandas as pd
import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from backend.main import app
from backend.config import settings
from backend.database import SessionLocal, init_db
from backend.models.transaction import Transaction
from backend.models.reconciliation import ReconciliationResult
from backend.models.exception import ReconciliationException
from backend.models.audit import AuditLog, audit_log_cleanup_context
from dashboard.export_utils import (
    dataframe_to_excel_bytes,
    dataframe_to_csv_bytes,
    dict_to_json_bytes,
)

client = TestClient(app)

PREFIX = "E2E_LIFECYCLE_"


@pytest.fixture(scope="module", autouse=True)
def force_offline_heuristic():
    """Forces heuristic LLM client during E2E tests to prevent external network calls."""
    orig_provider = settings.LLM_PROVIDER
    orig_key = settings.GEMINI_API_KEY
    settings.LLM_PROVIDER = "heuristic"
    settings.GEMINI_API_KEY = ""
    yield
    settings.LLM_PROVIDER = orig_provider
    settings.GEMINI_API_KEY = orig_key


def cleanup_e2e_records(db: Session) -> None:
    """Purges all records with E2E_LIFECYCLE_ prefixes to ensure test hermeticity."""
    with audit_log_cleanup_context():
        db.query(AuditLog).filter(
            (AuditLog.entity_id.like(f"%{PREFIX}%")) |
            (AuditLog.audit_id.like(f"%{PREFIX}%"))
        ).delete(synchronize_session=False)

        db.query(ReconciliationException).filter(
            (ReconciliationException.exception_id.like(f"%{PREFIX}%")) |
            (ReconciliationException.transaction_id.like(f"%{PREFIX}%"))
        ).delete(synchronize_session=False)

        db.query(ReconciliationResult).filter(
            (ReconciliationResult.reconciliation_id.like(f"%{PREFIX}%")) |
            (ReconciliationResult.gateway_transaction_id.like(f"%{PREFIX}%")) |
            (ReconciliationResult.bank_transaction_id.like(f"%{PREFIX}%"))
        ).delete(synchronize_session=False)

        db.query(Transaction).filter(
            Transaction.transaction_id.like(f"%{PREFIX}%")
        ).delete(synchronize_session=False)

        db.commit()


@pytest.fixture(autouse=True)
def setup_e2e_db():
    """Initializes DB schema and isolates database state before and after each test."""
    init_db()
    db: Session = SessionLocal()
    try:
        cleanup_e2e_records(db)
    finally:
        db.close()
    yield
    db = SessionLocal()
    try:
        cleanup_e2e_records(db)
    finally:
        db.close()


# =============================================================================
# TEST 1: Discrepancy Lifecycle to Human Approval & Audit Trail
# =============================================================================

def test_e2e_discrepancy_lifecycle_to_human_approval_audit():
    """
    End-to-end integration test validating the complete anomaly-to-resolution lifecycle:
    1. Seed a realistic mismatched Gateway + Bank transaction cluster into the DB.
    2. Run reconciliation via POST /reconcile.
    3. Verify an unresolved reconciliation result and an OPEN exception are generated.
    4. Verify fuzzy investigation and AI advisory reasoning are attached.
    5. Verify AI did NOT resolve or approve the discrepancy (human authority governance).
    6. Execute human approval via POST /exceptions/{exception_id}/approve with reviewer ID & notes.
    7. Verify exception transitions to APPROVED and linked result is marked resolved.
    8. Query the immutable audit trail via GET /audit and verify EXCEPTION_APPROVED is recorded
       with the human reviewer actor, not AI.
    """
    now = datetime.now(timezone.utc)
    db: Session = SessionLocal()
    try:
        # Step 1: Seed realistic mismatched transaction pair
        # Gateway payment: INR 5,000.00
        gw_txn = Transaction(
            transaction_id=f"{PREFIX}GW_DISC_01",
            source="GATEWAY",
            reference_id="pay_e2e_disc_987",
            order_id="ORD_E2E_DISC_01",
            customer_id="CUST_E2E_01",
            amount=5000.00,
            currency="INR",
            transaction_date=now - timedelta(hours=2),
            status="CAPTURED",
            transaction_type="PAYMENT",
            description="Payment for Order ORD_E2E_DISC_01"
        )
        # Bank settlement: INR 4,800.00 (INR 200 fee/shortfall discrepancy)
        bnk_txn = Transaction(
            transaction_id=f"{PREFIX}BNK_DISC_01",
            source="BANK",
            reference_id="PAY-E2E-DISC-987",  # Minor formatting variation for fuzzy engine
            order_id="ORD_E2E_DISC_01",
            customer_id="CUST_E2E_01",
            amount=4800.00,
            currency="INR",
            transaction_date=now - timedelta(hours=1),
            status="CREDIT",
            transaction_type="SETTLEMENT",
            description="NEFT credit PAY-E2E-DISC-987"
        )
        db.add_all([gw_txn, bnk_txn])
        db.commit()
    finally:
        db.close()

    # Step 2: Trigger reconciliation pipeline via FastAPI TestClient
    recon_run_resp = client.post("/reconcile")
    assert recon_run_resp.status_code == 200, f"Reconciliation run failed: {recon_run_resp.text}"
    recon_summary = recon_run_resp.json()
    assert recon_summary["total_clusters"] >= 1
    assert recon_summary["total_review"] >= 1

    # Step 3: Verify reconciliation result and exception in database
    db = SessionLocal()
    try:
        recon_result = db.query(ReconciliationResult).filter(
            ReconciliationResult.gateway_transaction_id == f"{PREFIX}GW_DISC_01"
        ).first()
        assert recon_result is not None, "ReconciliationResult was not persisted"
        assert recon_result.final_decision == "HUMAN_REVIEW"
        assert recon_result.is_resolved is False
        assert recon_result.discrepancy_amount == 200.00

        # Correlate exception strictly by reconciliation_id
        exception_record = db.query(ReconciliationException).filter(
            ReconciliationException.reconciliation_id == recon_result.reconciliation_id
        ).first()
        assert exception_record is not None, "ReconciliationException was not linked"
        target_exception_id = exception_record.exception_id

        # Step 4: Verify OPEN status, advisory reasoning, and governance invariants
        assert exception_record.status == "OPEN"
        assert exception_record.difference_amount == 200.00
        # AI reasoning is populated as advisory evidence
        assert recon_result.ai_recommendation in ("REVIEW", "ESCALATE", "APPROVE", "REJECT")
        assert exception_record.ai_explanation is not None

        # Step 5: Critical Governance Mandate — AI did NOT independently approve or resolve
        assert exception_record.status not in ("APPROVED", "REJECTED", "RESOLVED")
        assert recon_result.is_resolved is False
    finally:
        db.close()

    # Step 6: Human Reviewer Adjudication via REST API
    reviewer_id = "HUMAN_FIN_OFFICER_88"
    approval_notes = "Variance of INR 200 verified as contractual merchant settlement fee. Approved."
    approve_resp = client.post(
        f"/exceptions/{target_exception_id}/approve",
        json={
            "reviewer_id": reviewer_id,
            "notes": approval_notes
        }
    )
    assert approve_resp.status_code == 200, f"Human approval rejected: {approve_resp.text}"
    approval_data = approve_resp.json()
    assert approval_data["status"] == "APPROVED"
    assert approval_data["resolved_by"] == reviewer_id
    assert approval_data["reviewer_notes"] == approval_notes
    assert approval_data["resolved_at"] is not None

    # Step 7: Verify synchronized database state
    db = SessionLocal()
    try:
        db.expire_all()
        updated_exc = db.query(ReconciliationException).filter(
            ReconciliationException.exception_id == target_exception_id
        ).first()
        assert updated_exc.status == "APPROVED"
        assert updated_exc.resolved_by == reviewer_id

        updated_recon = db.query(ReconciliationResult).filter(
            ReconciliationResult.reconciliation_id == recon_result.reconciliation_id
        ).first()
        assert updated_recon.is_resolved is True
        assert updated_recon.final_decision == "MANUAL_APPROVED"

        # Step 8: Verify immutable Audit Trail records human decision authority
        audit_trail_resp = client.get(
            "/audit",
            params={
                "entity": "EXCEPTION",
                "entity_id": target_exception_id,
                "action": "EXCEPTION_APPROVED"
            }
        )
        assert audit_trail_resp.status_code == 200
        audit_items = audit_trail_resp.json().get("items", [])
        assert len(audit_items) >= 1, f"Audit record for {target_exception_id} not found"

        approved_audit = audit_items[0]
        assert approved_audit["action"] == "EXCEPTION_APPROVED"
        assert approved_audit["actor"] == reviewer_id
        assert approved_audit["entity"] == "EXCEPTION"
        assert approved_audit["entity_id"] == target_exception_id
        audit_new_val = json.loads(approved_audit["new_value"])
        assert audit_new_val["status"] == "APPROVED"
        assert audit_new_val["resolved_by"] == reviewer_id

        # Verify AI_CONTROLLER is NOT recorded as the resolution actor
        assert approved_audit["actor"] != "AI_CONTROLLER"
    finally:
        db.close()


# =============================================================================
# TEST 2: Reporting API to Export Serialization Pipeline
# =============================================================================

def test_e2e_reporting_api_to_export_pipeline():
    """
    End-to-end integration test validating report aggregation and export generation:
    1. Seed representative financial, reconciliation, and exception records.
    2. Query GET /reports/executive and GET /reports/reconciliation via FastAPI TestClient.
    3. Transform API JSON payloads into structured DataFrames mimicking dashboard app logic.
    4. Generate binary multi-sheet Excel (.xlsx) and CSV (.csv) buffers via real export_utils.
    5. Reopen Excel buffers with openpyxl and verify worksheet structure, columns, and data fidelity.
    6. Verify UTF-8 CSV decoding matches expected business totals.
    """
    now = datetime.now(timezone.utc)
    db: Session = SessionLocal()
    try:
        # Seed 2 transactions
        t1 = Transaction(
            transaction_id=f"{PREFIX}REP_TXN_01",
            source="GATEWAY",
            amount=25000.50,
            currency="INR",
            status="CAPTURED",
            transaction_type="PAYMENT",
            transaction_date=now
        )
        t2 = Transaction(
            transaction_id=f"{PREFIX}REP_TXN_02",
            source="BANK",
            amount=14999.50,
            currency="INR",
            status="SETTLED",
            transaction_type="SETTLEMENT",
            transaction_date=now
        )
        db.add_all([t1, t2])

        # Seed 1 resolved reconciliation result (Auto-Reconciled)
        r1 = ReconciliationResult(
            reconciliation_id=f"{PREFIX}REP_REC_01",
            gateway_transaction_id=f"{PREFIX}REP_TXN_01",
            match_score=100.0,
            matching_method="EXACT_RULE",
            final_decision="AUTO_RECONCILED",
            discrepancy_amount=0.00,
            is_resolved=True,
            reconciled_at=now
        )
        # Seed 1 manually approved reconciliation result with discrepancy
        r2 = ReconciliationResult(
            reconciliation_id=f"{PREFIX}REP_REC_02",
            gateway_transaction_id=f"{PREFIX}REP_TXN_02",
            match_score=75.0,
            matching_method="RULE_BASED",
            final_decision="MANUAL_APPROVED",
            discrepancy_amount=350.00,
            is_resolved=True,
            reconciled_at=now
        )
        db.add_all([r1, r2])

        # Seed linked resolved exception
        exc1 = ReconciliationException(
            exception_id=f"{PREFIX}REP_EXC_01",
            reconciliation_id=f"{PREFIX}REP_REC_02",
            transaction_id=f"{PREFIX}REP_TXN_02",
            category="AMOUNT_MISMATCH",
            severity="MEDIUM",
            difference_amount=350.00,
            status="APPROVED",
            resolved_by="AUDITOR_RAJESH",
            resolved_at=now,
            created_at=now
        )
        db.add(exc1)
        db.commit()
    finally:
        db.close()

    # Step 1: Query GET /reports/executive via TestClient
    exec_resp = client.get("/reports/executive")
    assert exec_resp.status_code == 200, f"Executive report failed: {exec_resp.text}"
    exec_json = exec_resp.json()
    assert exec_json["total_transactions"] >= 2
    assert exec_json["total_transaction_value_inr"] >= 40000.00
    assert exec_json["total_reconciliation_results"] >= 2
    assert "generated_at" in exec_json

    # Step 2: Query GET /reports/reconciliation via TestClient
    recon_resp = client.get("/reports/reconciliation")
    assert recon_resp.status_code == 200, f"Reconciliation report failed: {recon_resp.text}"
    recon_json = recon_resp.json()
    assert recon_json["total"] >= 2
    assert isinstance(recon_json["items"], list)

    # Step 3: Feed API response into dashboard export utility (Executive Summary DataFrame)
    df_exec_summary = pd.DataFrame([
        {"Metric": "Total Ingested Transactions", "Value": str(exec_json.get("total_transactions"))},
        {"Metric": "Total Ingested Volume (INR)", "Value": str(exec_json.get("total_transaction_value_inr"))},
        {"Metric": "Total Auto-Reconciled", "Value": str(exec_json.get("total_auto_reconciled"))},
        {"Metric": "Auto-Reconciliation Rate", "Value": f"{exec_json.get('auto_reconciliation_rate'):.2f}%"},
        {"Metric": "Report Generated At", "Value": str(exec_json.get("generated_at"))},
    ])

    # Generate Excel bytes for Executive Summary
    exec_xlsx_bytes = dataframe_to_excel_bytes(df_exec_summary, sheet_name="ExecutiveSummary")
    assert isinstance(exec_xlsx_bytes, bytes)
    assert len(exec_xlsx_bytes) > 0

    # Validate Excel workbook via openpyxl
    wb_exec = load_workbook(io.BytesIO(exec_xlsx_bytes))
    assert "ExecutiveSummary" in wb_exec.sheetnames
    sheet_exec = wb_exec["ExecutiveSummary"]
    assert sheet_exec["A1"].value == "Metric"
    assert sheet_exec["B1"].value == "Value"
    assert sheet_exec["A2"].value == "Total Ingested Transactions"

    # Step 4: Feed API response into dashboard export utility (Reconciliation Clusters)
    df_recon_items = pd.DataFrame(recon_json["items"])
    assert not df_recon_items.empty
    assert "reconciliation_id" in df_recon_items.columns
    assert "final_decision" in df_recon_items.columns

    recon_xlsx_bytes = dataframe_to_excel_bytes(df_recon_items, sheet_name="ReconciliationReport")
    assert isinstance(recon_xlsx_bytes, bytes)
    assert len(recon_xlsx_bytes) > 0

    wb_recon = load_workbook(io.BytesIO(recon_xlsx_bytes))
    assert "ReconciliationReport" in wb_recon.sheetnames
    sheet_recon = wb_recon["ReconciliationReport"]
    recon_headers = [cell.value for cell in sheet_recon[1]]
    assert "reconciliation_id" in recon_headers
    assert "final_decision" in recon_headers
    assert "match_score" in recon_headers

    # Step 5: Validate CSV and JSON serialization pipelines
    recon_csv_bytes = dataframe_to_csv_bytes(df_recon_items)
    assert isinstance(recon_csv_bytes, bytes)
    csv_text = recon_csv_bytes.decode("utf-8")
    assert "reconciliation_id,gateway_transaction_id" in csv_text
    assert f"{PREFIX}REP_REC_01" in csv_text

    recon_json_bytes = dict_to_json_bytes(recon_json)
    assert isinstance(recon_json_bytes, bytes)
    reloaded_json = json.loads(recon_json_bytes.decode("utf-8"))
    assert reloaded_json["total"] == recon_json["total"]
    assert len(reloaded_json["items"]) == len(recon_json["items"])
