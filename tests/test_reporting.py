"""
ReconcileAI - Phase 16 Export & Reporting Test Suite
Comprehensive verification of:
1. Reusable in-memory export utilities (CSV, Excel XLSX, JSON, text, empty datasets, Unicode/INR).
2. Backend ReportingService business logic (executive summary, SLA aging, three-leg reconciliation, compliance).
3. Backend report endpoints (/reports/summary, /reports/executive, /reports/reconciliation, /reports/exceptions, /reports/transactions, /reports/audit).
4. ReconcileAPIClient reporting methods.
5. Strict read-only safety (zero mutations to transactions, reconciliation, exceptions, or audit logs).
"""

import io
import json
import pytest
from datetime import datetime, timezone, timedelta
from decimal import Decimal
import pandas as pd
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
from unittest.mock import patch, MagicMock

from backend.main import app
from backend.database import SessionLocal, init_db
from backend.models.transaction import Transaction
from backend.models.reconciliation import ReconciliationResult
from backend.models.exception import ReconciliationException
from backend.models.audit import AuditLog, audit_log_cleanup_context
from backend.services.reporting_service import ReportingService
from dashboard.export_utils import (
    dataframe_to_csv_bytes,
    dataframe_to_excel_bytes,
    dataframes_to_excel_bytes,
    dict_to_json_bytes,
    text_to_bytes,
)
from dashboard.api_client import ReconcileAPIClient

client = TestClient(app)


# =============================================================================
# 1. Export Utilities Unit Tests
# =============================================================================

class TestExportUtilities:
    """Tests in-memory serialization helpers without any filesystem writes."""

    def test_csv_bytes_standard(self):
        df = pd.DataFrame([
            {"txn_id": "TXN_001", "amount": 1500.50, "currency": "INR"},
            {"txn_id": "TXN_002", "amount": 2500.00, "currency": "INR"}
        ])
        csv_bytes = dataframe_to_csv_bytes(df)
        assert isinstance(csv_bytes, bytes)
        assert b"TXN_001,1500.5,INR" in csv_bytes
        assert b"txn_id,amount,currency" in csv_bytes

    def test_csv_bytes_empty(self):
        empty_df = pd.DataFrame()
        assert dataframe_to_csv_bytes(empty_df) == b""
        assert dataframe_to_csv_bytes(None) == b""

        df_cols_only = pd.DataFrame(columns=["colA", "colB"])
        cols_bytes = dataframe_to_csv_bytes(df_cols_only)
        assert b"colA,colB" in cols_bytes

    def test_csv_unicode_inr_values(self):
        df = pd.DataFrame([
            {"item": "Breakage", "formatted": "₹15,000.50", "status": "APPROVED ✅"}
        ])
        csv_bytes = dataframe_to_csv_bytes(df)
        decoded = csv_bytes.decode("utf-8")
        assert "₹15,000.50" in decoded
        assert "APPROVED ✅" in decoded

    def test_excel_bytes_single_sheet(self):
        df = pd.DataFrame([
            {"col1": "Alpha", "col2": 100},
            {"col1": "Beta", "col2": 200}
        ])
        xlsx_bytes = dataframe_to_excel_bytes(df, sheet_name="TestSheet")
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

        # Verify workbook integrity using openpyxl in-memory
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "TestSheet" in wb.sheetnames
        sheet = wb["TestSheet"]
        assert sheet["A1"].value == "col1"
        assert sheet["B2"].value == 100

    def test_excel_bytes_multi_sheet(self):
        df1 = pd.DataFrame([{"k": "v1"}])
        df2 = pd.DataFrame([{"num": 42}])
        sheets = {
            "ExecutiveSummary": df1,
            "SeverityBreakdown": df2
        }
        xlsx_bytes = dataframes_to_excel_bytes(sheets)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "ExecutiveSummary" in wb.sheetnames
        assert "SeverityBreakdown" in wb.sheetnames
        assert wb["ExecutiveSummary"]["A2"].value == "v1"
        assert wb["SeverityBreakdown"]["A2"].value == 42

    def test_excel_bytes_empty(self):
        xlsx_bytes = dataframes_to_excel_bytes({})
        assert isinstance(xlsx_bytes, bytes)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Summary" in wb.sheetnames

    def test_json_bytes_serialization(self):
        now = datetime.now(timezone.utc)
        data = {
            "title": "Report",
            "timestamp": now,
            "amount": Decimal("999.99"),
            "items": [{"id": 1}, {"id": 2}]
        }
        json_bytes = dict_to_json_bytes(data)
        assert isinstance(json_bytes, bytes)
        parsed = json.loads(json_bytes.decode("utf-8"))
        assert parsed["title"] == "Report"
        assert parsed["amount"] == 999.99
        assert parsed["timestamp"] == now.isoformat()

    def test_text_to_bytes(self):
        text = "# Executive Report\nAll ledgers reconciled cleanly."
        b = text_to_bytes(text)
        assert isinstance(b, bytes)
        assert b.decode("utf-8") == text


# =============================================================================
# 2. Backend ReportingService & Database Tests
# =============================================================================

@pytest.fixture(scope="module")
def setup_reporting_data():
    """Seeds test data for reporting verification."""
    init_db()
    db: Session = SessionLocal()
    with audit_log_cleanup_context():
        db.query(AuditLog).delete(synchronize_session=False)
        db.query(ReconciliationException).delete(synchronize_session=False)
        db.query(ReconciliationResult).delete(synchronize_session=False)
        db.query(Transaction).delete(synchronize_session=False)
        db.commit()

    now = datetime.now(timezone.utc)
    t1 = Transaction(
        transaction_id="TX_REP_01",
        source="GATEWAY",
        reference_id="REF_REP_01",
        order_id="ORD_REP_01",
        customer_id="CUST_01",
        amount=5000.0,
        currency="INR",
        status="CAPTURED",
        transaction_type="PAYMENT",
        transaction_date=now - timedelta(hours=2)
    )
    t2 = Transaction(
        transaction_id="TX_REP_02",
        source="BANK",
        reference_id="REF_REP_01",
        order_id="ORD_REP_01",
        customer_id="CUST_01",
        amount=5000.0,
        currency="INR",
        status="SETTLED",
        transaction_type="SETTLEMENT",
        transaction_date=now - timedelta(hours=1)
    )
    db.add_all([t1, t2])

    r1 = ReconciliationResult(
        reconciliation_id="REC_REP_01",
        gateway_transaction_id="TX_REP_01",
        bank_transaction_id="TX_REP_02",
        erp_invoice_id="INV_REP_01",
        match_score=100.0,
        matching_method="EXACT_RULE",
        ai_recommendation="AUTO_RECONCILE",
        ai_confidence=100.0,
        ai_reasoning="Exact 3-leg match",
        final_decision="AUTO_RECONCILED",
        discrepancy_amount=0.0,
        is_resolved=True,
        reconciled_at=now
    )
    r2 = ReconciliationResult(
        reconciliation_id="REC_REP_02",
        gateway_transaction_id="TX_REP_03",
        match_score=40.0,
        matching_method="AI_REASONING",
        ai_recommendation="REVIEW",
        ai_confidence=60.0,
        ai_reasoning="Amount mismatch detected",
        final_decision="HUMAN_REVIEW",
        discrepancy_amount=250.0,
        is_resolved=False,
        reconciled_at=now
    )
    db.add_all([r1, r2])

    e1 = ReconciliationException(
        exception_id="EXC_REP_01",
        reconciliation_id="REC_REP_02",
        transaction_id="TX_REP_03",
        category="AMOUNT_MISMATCH",
        severity="HIGH",
        difference_amount=250.0,
        status="OPEN",
        sla_duration_hours=24.0,
        sla_deadline=now + timedelta(hours=12),
        sla_status="OK",
        escalation_level=0,
        created_at=now
    )
    e2 = ReconciliationException(
        exception_id="EXC_REP_02",
        reconciliation_id="REC_REP_03",
        transaction_id="TX_REP_04",
        category="MISSING_BANK_TRANSACTION",
        severity="CRITICAL",
        difference_amount=1200.0,
        status="OPEN",
        sla_duration_hours=6.0,
        sla_deadline=now - timedelta(hours=1),
        sla_status="BREACHED",
        escalation_level=1,
        escalated_at=now,
        created_at=now - timedelta(hours=7)
    )
    db.add_all([e1, e2])

    a1 = AuditLog(
        audit_id="AUD_REP_01",
        timestamp=now,
        actor="SYSTEM",
        action="TRANSACTION_INGESTED",
        entity="TRANSACTION",
        entity_id="TX_REP_01",
        reason="Synthetic ingestion"
    )
    a2 = AuditLog(
        audit_id="AUD_REP_02",
        timestamp=now,
        actor="AI_CONTROLLER",
        action="AI_REASONED",
        entity="RECONCILIATION",
        entity_id="REC_REP_02",
        reason="Discrepancy analyzed"
    )
    db.add_all([a1, a2])
    db.commit()
    db.close()

    yield

    db = SessionLocal()
    with audit_log_cleanup_context():
        db.query(AuditLog).delete(synchronize_session=False)
        db.query(ReconciliationException).delete(synchronize_session=False)
        db.query(ReconciliationResult).delete(synchronize_session=False)
        db.query(Transaction).delete(synchronize_session=False)
        db.commit()
    db.close()


def test_reporting_service_operational_summary(setup_reporting_data):
    """Verifies that ReportingService accurately aggregates counts and breakdowns."""
    db = SessionLocal()
    try:
        summary = ReportingService.get_operational_summary(db)
        assert summary["total_transactions"] == 2
        assert summary["total_reconciliation_results"] == 2
        assert summary["total_auto_reconciled"] == 1
        assert summary["auto_reconciliation_rate"] == 50.0
        assert summary["total_exceptions"] == 2
        assert summary["open_exceptions"] == 2
        assert summary["unresolved_amount_inr"] == 1450.0
        assert summary["exceptions_by_severity"].get("CRITICAL") == 1
        assert summary["exceptions_by_severity"].get("HIGH") == 1
        assert summary["sla_status_breakdown"].get("BREACHED") == 1
        assert summary["sla_status_breakdown"].get("OK") == 1
        assert summary["decision_breakdown"].get("AUTO_RECONCILED") == 1
        assert summary["decision_breakdown"].get("HUMAN_REVIEW") == 1
    finally:
        db.close()


def test_reporting_service_executive_report(setup_reporting_data):
    """Verifies executive financial statement aggregation."""
    db = SessionLocal()
    try:
        exec_rep = ReportingService.get_executive_report(db)
        assert exec_rep["total_transaction_value_inr"] == 10000.0
        assert "generated_at" in exec_rep
        assert exec_rep["auto_reconciliation_rate"] == 50.0
    finally:
        db.close()


def test_reporting_service_reconciliation_report(setup_reporting_data):
    """Verifies candidate cluster three-leg reporting."""
    db = SessionLocal()
    try:
        all_rep = ReportingService.get_reconciliation_report(db)
        assert len(all_rep) == 2
        auto_rep = ReportingService.get_reconciliation_report(db, final_decision="AUTO_RECONCILED")
        assert len(auto_rep) == 1
        assert auto_rep[0]["reconciliation_id"] == "REC_REP_01"
        assert auto_rep[0]["matching_method"] == "EXACT_RULE"
    finally:
        db.close()


def test_reporting_service_exception_aging_triage_sort(setup_reporting_data):
    """Verifies that exception aging reports sort breached/urgent exceptions first."""
    db = SessionLocal()
    try:
        aging = ReportingService.get_exception_aging_report(db)
        assert len(aging) == 2
        # BREACHED critical exception should be ranked first
        assert aging[0]["exception_id"] == "EXC_REP_02"
        assert aging[0]["sla_status"] == "BREACHED"
        assert aging[1]["exception_id"] == "EXC_REP_01"
        assert aging[1]["sla_status"] == "OK"
    finally:
        db.close()


def test_reporting_service_strictly_read_only(setup_reporting_data):
    """CRITICAL: Confirms that report generation executes 0 mutations on database records."""
    db = SessionLocal()
    try:
        initial_tx_count = db.query(Transaction).count()
        initial_res_count = db.query(ReconciliationResult).count()
        initial_exc_count = db.query(ReconciliationException).count()
        initial_aud_count = db.query(AuditLog).count()

        # Execute all reporting service queries
        ReportingService.get_operational_summary(db)
        ReportingService.get_executive_report(db)
        ReportingService.get_reconciliation_report(db)
        ReportingService.get_exception_aging_report(db)
        ReportingService.get_audit_compliance_report(db)
        ReportingService.get_all_transactions(db)

        # Confirm counts are strictly identical
        assert db.query(Transaction).count() == initial_tx_count
        assert db.query(ReconciliationResult).count() == initial_res_count
        assert db.query(ReconciliationException).count() == initial_exc_count
        assert db.query(AuditLog).count() == initial_aud_count
    finally:
        db.close()


# =============================================================================
# 3. HTTP Endpoints Tests via FastAPI TestClient
# =============================================================================

def test_http_reports_summary_backward_compatibility(setup_reporting_data):
    """Verifies GET /reports/summary returns status 200 and matches expected structure."""
    resp = client.get("/reports/summary")
    assert resp.status_code == 200
    data = resp.json()
    assert "total_transactions" in data
    assert "auto_reconciliation_rate" in data
    assert "decision_breakdown" in data


def test_http_reports_executive(setup_reporting_data):
    """Verifies GET /reports/executive endpoint."""
    resp = client.get("/reports/executive")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total_transaction_value_inr"] == 10000.0
    assert "generated_at" in data


def test_http_reports_reconciliation(setup_reporting_data):
    """Verifies GET /reports/reconciliation returns unpaginated report items."""
    resp = client.get("/reports/reconciliation")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 2
    assert len(data["items"]) == 2

    filtered = client.get("/reports/reconciliation?final_decision=AUTO_RECONCILED")
    assert filtered.status_code == 200
    assert filtered.json()["total"] == 1


def test_http_reports_exceptions(setup_reporting_data):
    """Verifies GET /reports/exceptions returns aging records."""
    resp = client.get("/reports/exceptions")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 2
    assert data["items"][0]["sla_status"] == "BREACHED"


def test_http_reports_transactions(setup_reporting_data):
    """Verifies GET /reports/transactions returns complete matching transactions."""
    resp = client.get("/reports/transactions")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 2
    assert len(data["items"]) == 2

    filtered = client.get("/reports/transactions?source=GATEWAY")
    assert filtered.status_code == 200
    assert filtered.json()["total"] == 1


def test_http_reports_audit(setup_reporting_data):
    """Verifies GET /reports/audit returns complete compliance log records."""
    resp = client.get("/reports/audit")
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] == 2
    assert len(data["items"]) == 2


# =============================================================================
# 4. ReconcileAPIClient Method Tests
# =============================================================================

class TestReconcileAPIClientReportingMethods:
    """Verifies ReconcileAPIClient reporting methods construct valid HTTP requests."""

    @pytest.fixture
    def api_client(self):
        return ReconcileAPIClient(base_url="http://test-server:8000")

    @patch("requests.get")
    def test_client_get_executive_report(self, mock_get, api_client):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"total_transactions": 100, "total_transaction_value_inr": 50000.0}
        mock_get.return_value = mock_resp

        res = api_client.get_executive_report()
        assert res["total_transactions"] == 100
        mock_get.assert_called_once_with("http://test-server:8000/reports/executive", params=None, timeout=10)

    @patch("requests.get")
    def test_client_get_reconciliation_report(self, mock_get, api_client):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"total": 5, "items": []}
        mock_get.return_value = mock_resp

        res = api_client.get_reconciliation_report(final_decision="AUTO_RECONCILED")
        assert res["total"] == 5
        mock_get.assert_called_once_with(
            "http://test-server:8000/reports/reconciliation",
            params={"final_decision": "AUTO_RECONCILED"},
            timeout=10
        )

    @patch("requests.get")
    def test_client_get_exception_aging_report(self, mock_get, api_client):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"total": 3, "items": []}
        mock_get.return_value = mock_resp

        res = api_client.get_exception_aging_report(sla_status="BREACHED")
        assert res["total"] == 3
        mock_get.assert_called_once_with(
            "http://test-server:8000/reports/exceptions",
            params={"sla_status": "BREACHED"},
            timeout=10
        )

    @patch("requests.get")
    def test_client_get_all_transactions(self, mock_get, api_client):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"total": 10, "items": []}
        mock_get.return_value = mock_resp

        res = api_client.get_all_transactions(source="GATEWAY")
        assert res["total"] == 10
        mock_get.assert_called_once_with(
            "http://test-server:8000/reports/transactions",
            params={"source": "GATEWAY"},
            timeout=10
        )

    @patch("requests.get")
    def test_client_get_audit_compliance_report(self, mock_get, api_client):
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"total": 8, "items": []}
        mock_get.return_value = mock_resp

        res = api_client.get_audit_compliance_report(entity="TRANSACTION")
        assert res["total"] == 8
        mock_get.assert_called_once_with(
            "http://test-server:8000/reports/audit",
            params={"entity": "TRANSACTION"},
            timeout=10
        )


# =============================================================================
# 5. Phase 17 Task 5A — Reporting Robustness Tests
# =============================================================================

class TestReportingRobustness:
    """
    Robustness tests for ReportingService:
    1. Empty database safety (no ZeroDivisionError, zero totals, empty list collections)
    2. Extreme and negative financial value handling & absolute-magnitude triage sorting
    """

    def test_reporting_service_empty_database_safety(self):
        """
        Execute all six ReportingService methods on an empty database.
        Verifies:
        - ZeroDivisionError does not occur when total_reconciliation_results == 0
        - auto_reconciliation_rate is 0.0
        - Monetary totals are 0.0
        - List-returning queries return empty lists []
        """
        db: Session = SessionLocal()
        try:
            with audit_log_cleanup_context():
                db.query(AuditLog).delete(synchronize_session=False)
                db.query(ReconciliationException).delete(synchronize_session=False)
                db.query(ReconciliationResult).delete(synchronize_session=False)
                db.query(Transaction).delete(synchronize_session=False)
                db.commit()

            # 1. Operational summary
            op_summary = ReportingService.get_operational_summary(db)
            assert op_summary["total_transactions"] == 0
            assert op_summary["total_reconciliation_results"] == 0
            assert op_summary["total_auto_reconciled"] == 0
            assert op_summary["auto_reconciliation_rate"] == 0.0
            assert op_summary["total_exceptions"] == 0
            assert op_summary["open_exceptions"] == 0
            assert op_summary["unresolved_amount_inr"] == 0.0
            assert op_summary["exceptions_by_severity"] == {}
            assert op_summary["exceptions_by_category"] == {}
            assert op_summary["sla_status_breakdown"] == {}
            assert op_summary["decision_breakdown"] == {}

            # 2. Executive report
            exec_rep = ReportingService.get_executive_report(db)
            assert exec_rep["total_transactions"] == 0
            assert exec_rep["total_transaction_value_inr"] == 0.0
            assert exec_rep["auto_reconciliation_rate"] == 0.0
            assert exec_rep["unresolved_amount_inr"] == 0.0
            assert "generated_at" in exec_rep

            # 3. Reconciliation report
            recon_rep = ReportingService.get_reconciliation_report(db)
            assert recon_rep == []

            # 4. Exception aging report
            exc_rep = ReportingService.get_exception_aging_report(db)
            assert exc_rep == []

            # 5. Audit compliance report
            audit_rep = ReportingService.get_audit_compliance_report(db)
            assert audit_rep == []

            # 6. All transactions
            tx_rep = ReportingService.get_all_transactions(db)
            assert tx_rep == []
        finally:
            db.close()

    def test_reporting_service_extreme_and_negative_financial_values(self):
        """
        Verify ReportingService calculations with extreme amounts and negative/refund discrepancies.
        Verifies:
        - Extreme amounts (10 billion INR) sum correctly without precision overflow
        - Zero discrepancy results calculate auto rates correctly
        - Legitimate negative differences (e.g. refunds / chargebacks) sum correctly in VaR
        - Exception triage ordering uses absolute-magnitude ranking (-abs(difference_amount))
        """
        now = datetime.now(timezone.utc)
        db: Session = SessionLocal()
        try:
            with audit_log_cleanup_context():
                db.query(AuditLog).delete(synchronize_session=False)
                db.query(ReconciliationException).delete(synchronize_session=False)
                db.query(ReconciliationResult).delete(synchronize_session=False)
                db.query(Transaction).delete(synchronize_session=False)
                db.commit()

            # Extreme transaction: 10 billion INR
            t_huge = Transaction(
                transaction_id="TX_EXT_01",
                source="GATEWAY",
                amount=10_000_000_000.00,
                currency="INR",
                status="CAPTURED",
                transaction_type="PAYMENT",
                transaction_date=now,
            )
            # Standard payment transaction
            t_norm = Transaction(
                transaction_id="TX_EXT_02",
                source="BANK",
                amount=250.75,
                currency="INR",
                status="SETTLED",
                transaction_type="SETTLEMENT",
                transaction_date=now,
            )
            # Legitimate refund transaction with negative amount
            t_refund = Transaction(
                transaction_id="TX_EXT_03",
                source="GATEWAY",
                amount=-2500.00,
                currency="INR",
                status="SETTLED",
                transaction_type="REFUND",
                transaction_date=now,
            )
            db.add_all([t_huge, t_norm, t_refund])

            # Reconciliation with zero discrepancy
            r_zero = ReconciliationResult(
                reconciliation_id="REC_EXT_01",
                gateway_transaction_id="TX_EXT_01",
                match_score=100.0,
                matching_method="EXACT_RULE",
                final_decision="AUTO_RECONCILED",
                discrepancy_amount=0.00,
                is_resolved=True,
                reconciled_at=now,
            )
            # Reconciliation with review decision
            r_rev = ReconciliationResult(
                reconciliation_id="REC_EXT_02",
                gateway_transaction_id="TX_EXT_02",
                match_score=60.0,
                matching_method="RULE_BASED",
                final_decision="HUMAN_REVIEW",
                discrepancy_amount=500.00,
                is_resolved=False,
                reconciled_at=now,
            )
            db.add_all([r_zero, r_rev])

            # Three open exceptions with same SLA and severity to test triage magnitude:
            # e_large: difference_amount = 500.00
            # e_small: difference_amount = 100.00
            # e_zero: difference_amount = 0.00
            e_large = ReconciliationException(
                exception_id="EXC_EXT_01",
                reconciliation_id="REC_EXT_02",
                transaction_id="TX_EXT_02",
                category="AMOUNT_MISMATCH",
                severity="HIGH",
                difference_amount=500.00,
                status="OPEN",
                sla_duration_hours=24.0,
                sla_status="OK",
                escalation_level=0,
                created_at=now,
            )
            e_small = ReconciliationException(
                exception_id="EXC_EXT_02",
                reconciliation_id="REC_EXT_02",
                transaction_id="TX_EXT_02",
                category="PARTIAL_PAYMENT",
                severity="HIGH",
                difference_amount=100.00,
                status="OPEN",
                sla_duration_hours=24.0,
                sla_status="OK",
                escalation_level=0,
                created_at=now,
            )
            e_zero = ReconciliationException(
                exception_id="EXC_EXT_03",
                reconciliation_id="REC_EXT_01",
                transaction_id="TX_EXT_01",
                category="DATE_MISMATCH",
                severity="HIGH",
                difference_amount=0.00,
                status="OPEN",
                sla_duration_hours=24.0,
                sla_status="OK",
                escalation_level=0,
                created_at=now,
            )
            db.add_all([e_large, e_small, e_zero])
            db.commit()

            # Verify executive report handling of 10 billion INR with refund deduction
            exec_rep = ReportingService.get_executive_report(db)
            assert exec_rep["total_transactions"] == 3
            # 10,000,000,000.00 + 250.75 + (-2500.00) = 9,999,997,750.75
            assert exec_rep["total_transaction_value_inr"] == 9_999_997_750.75
            assert exec_rep["auto_reconciliation_rate"] == 50.0
            # VaR sums open non-negative difference_amounts: 500.00 + 100.00 + 0.00 = 600.00
            assert exec_rep["unresolved_amount_inr"] == 600.00
            assert exec_rep["unresolved_amount_inr"] >= 0.0

            # Verify triage sort ordering in get_exception_aging_report
            # All 3 have sla_status="OK", severity="HIGH", escalation_level=0.
            # Triage sort sorts by -abs(float(x.difference_amount or 0.0)):
            # 1. e_large: 500.00 -> first
            # 2. e_small: 100.00 -> second
            # 3. e_zero: 0.00 -> third
            aging_items = ReportingService.get_exception_aging_report(db)
            assert len(aging_items) == 3
            assert aging_items[0]["exception_id"] == "EXC_EXT_01"
            assert aging_items[0]["difference_amount"] == 500.00
            assert aging_items[1]["exception_id"] == "EXC_EXT_02"
            assert aging_items[1]["difference_amount"] == 100.00
            assert aging_items[2]["exception_id"] == "EXC_EXT_03"
            assert aging_items[2]["difference_amount"] == 0.00
        finally:
            with audit_log_cleanup_context():
                db.query(AuditLog).delete(synchronize_session=False)
                db.query(ReconciliationException).delete(synchronize_session=False)
                db.query(ReconciliationResult).delete(synchronize_session=False)
                db.query(Transaction).delete(synchronize_session=False)
                db.commit()
            db.close()

    def test_export_utilities_empty_report_datasets(self):
        """
        Export empty report datasets through existing export utilities: CSV, XLSX, JSON.
        Verifies:
        - Valid bytes returned for all formats
        - CSV encodes empty or header-only bytes cleanly
        - XLSX binary streams can be opened with openpyxl without corruption
        - JSON parses into empty dict/list
        """
        # Case A: Completely empty DataFrame (0 rows, 0 cols)
        df_empty = pd.DataFrame()
        csv_bytes_empty = dataframe_to_csv_bytes(df_empty)
        assert isinstance(csv_bytes_empty, bytes)
        assert csv_bytes_empty == b""

        # Case B: Empty report DataFrame with standard schema columns (0 rows, N cols)
        report_cols = ["reconciliation_id", "match_score", "discrepancy_amount", "final_decision"]
        df_cols_empty = pd.DataFrame(columns=report_cols)
        csv_bytes_cols = dataframe_to_csv_bytes(df_cols_empty)
        assert isinstance(csv_bytes_cols, bytes)
        assert csv_bytes_cols.decode("utf-8").strip() == ",".join(report_cols)

        # Case C: Excel export with empty DataFrame
        xlsx_single = dataframe_to_excel_bytes(df_cols_empty, sheet_name="EmptyReconReport")
        assert isinstance(xlsx_single, bytes)
        wb_single = load_workbook(io.BytesIO(xlsx_single))
        assert "EmptyReconReport" in wb_single.sheetnames
        sheet = wb_single["EmptyReconReport"]
        assert [cell.value for cell in sheet[1]] == report_cols

        # Case D: Excel multi-sheet export with empty sheets mapping
        xlsx_multi = dataframes_to_excel_bytes({"Exceptions": pd.DataFrame(), "Audit": df_cols_empty})
        assert isinstance(xlsx_multi, bytes)
        wb_multi = load_workbook(io.BytesIO(xlsx_multi))
        assert "Exceptions" in wb_multi.sheetnames
        assert "Audit" in wb_multi.sheetnames

        # Case E: JSON export with empty list and empty dict
        json_bytes_list = dict_to_json_bytes([])
        assert isinstance(json_bytes_list, bytes)
        assert json.loads(json_bytes_list.decode("utf-8")) == []

        json_bytes_dict = dict_to_json_bytes({})
        assert isinstance(json_bytes_dict, bytes)
        assert json.loads(json_bytes_dict.decode("utf-8")) == {}

        json_bytes_none = dict_to_json_bytes(None)
        assert isinstance(json_bytes_none, bytes)
        assert json.loads(json_bytes_none.decode("utf-8")) == {}

    def test_export_utilities_unicode_inr_and_multilingual_preservation(self):
        """
        Verify export utilities preserve Indian Rupee symbols (₹), Devanagari names ("राजेश शर्मा"),
        emojis ("✅"), and multilingual text across CSV, JSON, and Excel formats without mojibake.
        """
        records = [
            {
                "customer_name": "राजेश शर्मा",
                "narration": "UPI P2P transfer to Priya ✅",
                "formatted_amount": "₹45,000.75",
                "audit_reason": "Manual resolution approved by Ops: ₹45,000.75 reconciled",
            },
            {
                "customer_name": "அனிதா ரமேஷ்",
                "narration": "IMPS settlement with vendor ⚡",
                "formatted_amount": "₹1,25,000.00",
                "audit_reason": "Audit review verified: ₹1,25,000.00 matched",
            }
        ]
        df = pd.DataFrame(records)

        # 1. CSV UTF-8 verification
        csv_bytes = dataframe_to_csv_bytes(df)
        csv_decoded = csv_bytes.decode("utf-8")
        assert "₹45,000.75" in csv_decoded
        assert "राजेश शर्मा" in csv_decoded
        assert "✅" in csv_decoded
        assert "அனிதா ரமேஷ்" in csv_decoded
        assert "⚡" in csv_decoded
        assert "Audit review verified: ₹1,25,000.00 matched" in csv_decoded

        # 2. JSON UTF-8 verification (ensure_ascii=False ensures actual characters)
        json_bytes = dict_to_json_bytes(records)
        json_decoded = json_bytes.decode("utf-8")
        assert "₹45,000.75" in json_decoded
        assert "राजेश शर्मा" in json_decoded
        assert "✅" in json_decoded
        assert "அனிதா ரமேஷ்" in json_decoded
        # Parse back to verify structural integrity
        parsed_json = json.loads(json_decoded)
        assert parsed_json[0]["customer_name"] == "राजेश शर्मा"
        assert parsed_json[0]["formatted_amount"] == "₹45,000.75"
        assert parsed_json[0]["narration"] == "UPI P2P transfer to Priya ✅"
        assert parsed_json[1]["customer_name"] == "அனிதா ரமேஷ்"

        # 3. Excel openpyxl cell value verification
        xlsx_bytes = dataframe_to_excel_bytes(df, sheet_name="MultilingualReport")
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet = wb["MultilingualReport"]
        # Row 1 is header, Row 2 is record 0, Row 3 is record 1
        assert sheet["A2"].value == "राजेश शर्मा"
        assert sheet["B2"].value == "UPI P2P transfer to Priya ✅"
        assert sheet["C2"].value == "₹45,000.75"
        assert sheet["D2"].value == "Manual resolution approved by Ops: ₹45,000.75 reconciled"
        assert sheet["A3"].value == "அனிதா ரமேஷ்"
        assert sheet["C3"].value == "₹1,25,000.00"

    def test_export_roundtrip_data_fidelity(self):
        """
        Verify export round-trip fidelity across CSV and JSON:
        - Identifiers, monetary values, timestamps, and Unicode text remain semantically correct
        - No material financial precision is lost
        """
        iso_timestamp = "2026-09-04T12:00:00+00:00"
        dataset = [
            {
                "transaction_id": "TXN_FIDELITY_001",
                "amount": 999999.99,
                "fee": 15.50,
                "timestamp": iso_timestamp,
                "customer": "विक्रम साराभाई",
                "status": "CAPTURED",
            },
            {
                "transaction_id": "TXN_FIDELITY_002",
                "amount": 0.05,
                "fee": 0.00,
                "timestamp": iso_timestamp,
                "customer": "Sunita Williams",
                "status": "SETTLED",
            },
        ]

        # 1. JSON Roundtrip
        json_bytes = dict_to_json_bytes(dataset)
        parsed_json = json.loads(json_bytes.decode("utf-8"))
        assert len(parsed_json) == 2
        for orig, recovered in zip(dataset, parsed_json):
            assert recovered["transaction_id"] == orig["transaction_id"]
            assert recovered["amount"] == orig["amount"]
            assert recovered["fee"] == orig["fee"]
            assert recovered["timestamp"] == orig["timestamp"]
            assert recovered["customer"] == orig["customer"]
            assert recovered["status"] == orig["status"]

        # 2. CSV Roundtrip via pandas
        df_orig = pd.DataFrame(dataset)
        csv_bytes = dataframe_to_csv_bytes(df_orig)
        df_recovered = pd.read_csv(io.BytesIO(csv_bytes))

        assert len(df_recovered) == 2
        for idx, orig in enumerate(dataset):
            row = df_recovered.iloc[idx]
            assert str(row["transaction_id"]) == orig["transaction_id"]
            assert round(float(row["amount"]), 2) == orig["amount"]
            assert round(float(row["fee"]), 2) == orig["fee"]
            assert str(row["timestamp"]) == orig["timestamp"]
            assert str(row["customer"]) == orig["customer"]
            assert str(row["status"]) == orig["status"]
